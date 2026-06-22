"""
app.py - Main Flask Application
Employee Database Management System (EDMS)

Features:
  - Secure login/authentication (hashed passwords, sessions)
  - Role-based access (admin / hr / viewer)
  - Full CRUD for employees
  - Department management
  - Search & filter (by name, department, status)
  - Dashboard with live charts (Chart.js)
  - Attendance tracking
  - Salary slip generation (PDF)
  - Export employee list to Excel / PDF
  - Activity log (audit trail)
  - Photo upload for employee profile
"""

import io
import os
from datetime import datetime, date

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, send_file, abort
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from config import Config
import db as dbmod

# Excel / PDF export libraries
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# --------------------------------------------------------------------------
# App initialization
# --------------------------------------------------------------------------
app = Flask(__name__)
app.config.from_object(Config)
dbmod.init_app(app)

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)




def allowed_file(filename):
    return "." in filename and \
        filename.rsplit(".", 1)[1].lower() in app.config["ALLOWED_EXTENSIONS"]


def log_activity(action):
    """Write an entry into the activity_log table for audit purposes."""
    try:
        dbmod.execute_query(
            "INSERT INTO activity_log (user_id, action) VALUES (%s, %s)",
            (session.get("user_id"), action),
            commit=True,
        )
    except Exception:
        pass  # logging should never break the main flow


def login_required(role=None):
    """Decorator-factory to protect routes. role can be a list of allowed roles."""
    def decorator(f):
        from functools import wraps

        @wraps(f)
        def wrapped(*args, **kwargs):
            if "user_id" not in session:
                flash("Please login to continue.", "warning")
                return redirect(url_for("login"))
            if role and session.get("role") not in role:
                flash("You do not have permission to access that page.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return wrapped
    return decorator


# ==========================================================================
# AUTH ROUTES
# ==========================================================================

@app.route("/", methods=["GET"])
def index():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = dbmod.execute_query(
            "SELECT * FROM users WHERE username = %s", (username,), fetchone=True
        )

        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["user_id"]
            session["username"] = user["username"]
            session["full_name"] = user["full_name"]
            session["role"] = user["role"]

            dbmod.execute_query(
                "UPDATE users SET last_login = %s WHERE user_id = %s",
                (datetime.now(), user["user_id"]), commit=True
            )
            log_activity(f"User '{username}' logged in.")
            flash(f"Welcome back, {user['full_name']}!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password.", "danger")

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    """First-time admin setup / new HR account creation."""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip()
        role = request.form.get("role", "hr")

        existing = dbmod.execute_query(
            "SELECT user_id FROM users WHERE username=%s", (username,), fetchone=True
        )
        if existing:
            flash("Username already exists.", "danger")
            return render_template("register.html")

        pw_hash = generate_password_hash(password)
        dbmod.execute_query(
            """INSERT INTO users (username, password_hash, full_name, role, email)
               VALUES (%s, %s, %s, %s, %s)""",
            (username, pw_hash, full_name, role, email), commit=True
        )
        flash("Account created successfully. Please login.", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


@app.route("/logout")
def logout():
    log_activity(f"User '{session.get('username')}' logged out.")
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


# ==========================================================================
# DASHBOARD
# ==========================================================================

@app.route("/dashboard")
@login_required()
def dashboard():
    total_employees = dbmod.execute_query(
        "SELECT COUNT(*) AS c FROM employees WHERE status='Active'", fetchone=True
    )["c"]

    total_departments = dbmod.execute_query(
        "SELECT COUNT(*) AS c FROM departments", fetchone=True
    )["c"]

    total_payroll = dbmod.execute_query(
        "SELECT COALESCE(SUM(basic_salary+allowance-deduction),0) AS s "
        "FROM employees WHERE status='Active'", fetchone=True
    )["s"]

    new_this_month = dbmod.execute_query(
        "SELECT COUNT(*) AS c FROM employees "
        "WHERE MONTH(date_of_joining)=MONTH(CURDATE()) AND YEAR(date_of_joining)=YEAR(CURDATE())",
        fetchone=True
    )["c"]

    dept_stats = dbmod.execute_query(
        "SELECT * FROM department_stats", fetchall=True
    )

    recent_employees = dbmod.execute_query(
        "SELECT * FROM employee_full_view ORDER BY emp_id DESC LIMIT 5", fetchall=True
    )

    gender_stats = dbmod.execute_query(
        "SELECT gender, COUNT(*) AS c FROM employees WHERE status='Active' GROUP BY gender",
        fetchall=True
    )

    recent_logs = dbmod.execute_query(
        """SELECT al.action, al.log_time, u.full_name
           FROM activity_log al LEFT JOIN users u ON al.user_id = u.user_id
           ORDER BY al.log_id DESC LIMIT 6""",
        fetchall=True
    )

    return render_template(
        "dashboard.html",
        total_employees=total_employees,
        total_departments=total_departments,
        total_payroll=total_payroll,
        new_this_month=new_this_month,
        dept_stats=dept_stats,
        recent_employees=recent_employees,
        gender_stats=gender_stats,
        recent_logs=recent_logs,
    )


# ==========================================================================
# EMPLOYEE CRUD
# ==========================================================================

@app.route("/employees")
@login_required()
def employee_list():
    search = request.args.get("q", "").strip()
    dept_filter = request.args.get("dept", "")
    status_filter = request.args.get("status", "")
    page = int(request.args.get("page", 1))
    per_page = app.config["EMPLOYEES_PER_PAGE"]
    offset = (page - 1) * per_page

    query = "SELECT * FROM employee_full_view WHERE 1=1"
    params = []

    if search:
        query += " AND (full_name LIKE %s OR emp_code LIKE %s OR email LIKE %s)"
        like = f"%{search}%"
        params += [like, like, like]

    if dept_filter:
        query += " AND dept_name = %s"
        params.append(dept_filter)

    if status_filter:
        query += " AND status = %s"
        params.append(status_filter)

    count_query = query.replace("SELECT *", "SELECT COUNT(*) AS c")
    total = dbmod.execute_query(count_query, tuple(params), fetchone=True)["c"]

    query += " ORDER BY emp_id DESC LIMIT %s OFFSET %s"
    params += [per_page, offset]

    employees = dbmod.execute_query(query, tuple(params), fetchall=True)
    departments = dbmod.execute_query("SELECT * FROM departments ORDER BY dept_name", fetchall=True)

    total_pages = max(1, (total + per_page - 1) // per_page)

    return render_template(
        "employees.html",
        employees=employees,
        departments=departments,
        search=search,
        dept_filter=dept_filter,
        status_filter=status_filter,
        page=page,
        total_pages=total_pages,
        total=total,
    )


@app.route("/employees/add", methods=["GET", "POST"])
@login_required(role=["admin", "hr"])
def employee_add():
    departments = dbmod.execute_query("SELECT * FROM departments ORDER BY dept_name", fetchall=True)

    if request.method == "POST":
        form = request.form
        photo_filename = "default.png"

        file = request.files.get("photo")
        if file and file.filename and allowed_file(file.filename):
            photo_filename = secure_filename(
                f"{form.get('emp_code')}_{file.filename}"
            )
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], photo_filename))

        try:
            dbmod.execute_query(
                """INSERT INTO employees
                   (emp_code, first_name, last_name, email, phone, gender, dob, address,
                    dept_id, designation, date_of_joining, basic_salary, allowance,
                    deduction, photo, status)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (
                    form.get("emp_code"), form.get("first_name"), form.get("last_name"),
                    form.get("email"), form.get("phone"), form.get("gender"),
                    form.get("dob") or None, form.get("address"),
                    form.get("dept_id") or None, form.get("designation"),
                    form.get("date_of_joining") or None,
                    form.get("basic_salary") or 0, form.get("allowance") or 0,
                    form.get("deduction") or 0, photo_filename,
                    form.get("status", "Active"),
                ),
                commit=True,
            )
            log_activity(f"Added employee '{form.get('first_name')} {form.get('last_name')}'.")
            flash("Employee added successfully!", "success")
            return redirect(url_for("employee_list"))
        except Exception as e:
            flash(f"Error adding employee: {e}", "danger")

    return render_template("employee_form.html", departments=departments, employee=None)


@app.route("/employees/edit/<int:emp_id>", methods=["GET", "POST"])
@login_required(role=["admin", "hr"])
def employee_edit(emp_id):
    employee = dbmod.execute_query(
        "SELECT * FROM employees WHERE emp_id=%s", (emp_id,), fetchone=True
    )
    if not employee:
        abort(404)

    departments = dbmod.execute_query("SELECT * FROM departments ORDER BY dept_name", fetchall=True)

    if request.method == "POST":
        form = request.form
        photo_filename = employee["photo"]

        file = request.files.get("photo")
        if file and file.filename and allowed_file(file.filename):
            photo_filename = secure_filename(f"{form.get('emp_code')}_{file.filename}")
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], photo_filename))

        dbmod.execute_query(
            """UPDATE employees SET
                 first_name=%s, last_name=%s, email=%s, phone=%s, gender=%s, dob=%s,
                 address=%s, dept_id=%s, designation=%s, date_of_joining=%s,
                 basic_salary=%s, allowance=%s, deduction=%s, photo=%s, status=%s
               WHERE emp_id=%s""",
            (
                form.get("first_name"), form.get("last_name"), form.get("email"),
                form.get("phone"), form.get("gender"), form.get("dob") or None,
                form.get("address"), form.get("dept_id") or None, form.get("designation"),
                form.get("date_of_joining") or None, form.get("basic_salary") or 0,
                form.get("allowance") or 0, form.get("deduction") or 0,
                photo_filename, form.get("status", "Active"), emp_id,
            ),
            commit=True,
        )
        log_activity(f"Updated employee ID {emp_id}.")
        flash("Employee updated successfully!", "success")
        return redirect(url_for("employee_list"))

    return render_template("employee_form.html", departments=departments, employee=employee)


@app.route("/employees/delete/<int:emp_id>", methods=["POST"])
@login_required(role=["admin"])
def employee_delete(emp_id):
    dbmod.execute_query("DELETE FROM employees WHERE emp_id=%s", (emp_id,), commit=True)
    log_activity(f"Deleted employee ID {emp_id}.")
    flash("Employee deleted.", "info")
    return redirect(url_for("employee_list"))


@app.route("/employees/view/<int:emp_id>")
@login_required()
def employee_view(emp_id):
    employee = dbmod.execute_query(
        "SELECT * FROM employee_full_view WHERE emp_id=%s", (emp_id,), fetchone=True
    )
    if not employee:
        abort(404)

    attendance = dbmod.execute_query(
        "SELECT * FROM attendance WHERE emp_id=%s ORDER BY att_date DESC LIMIT 10",
        (emp_id,), fetchall=True
    )
    salary_history = dbmod.execute_query(
        "SELECT * FROM salary_history WHERE emp_id=%s ORDER BY salary_id DESC LIMIT 10",
        (emp_id,), fetchall=True
    )

    return render_template(
        "employee_view.html", employee=employee,
        attendance=attendance, salary_history=salary_history
    )


# ==========================================================================
# DEPARTMENTS
# ==========================================================================

@app.route("/departments", methods=["GET", "POST"])
@login_required(role=["admin", "hr"])
def departments():
    if request.method == "POST":
        dbmod.execute_query(
            "INSERT INTO departments (dept_name, dept_head, description) VALUES (%s,%s,%s)",
            (request.form.get("dept_name"), request.form.get("dept_head"),
             request.form.get("description")),
            commit=True,
        )
        flash("Department added.", "success")
        return redirect(url_for("departments"))

    dept_list = dbmod.execute_query("SELECT * FROM department_stats", fetchall=True)
    return render_template("departments.html", departments=dept_list)


@app.route("/departments/delete/<int:dept_id>", methods=["POST"])
@login_required(role=["admin"])
def department_delete(dept_id):
    dbmod.execute_query("DELETE FROM departments WHERE dept_id=%s", (dept_id,), commit=True)
    flash("Department deleted.", "info")
    return redirect(url_for("departments"))


# ==========================================================================
# ATTENDANCE
# ==========================================================================

@app.route("/attendance", methods=["GET", "POST"])
@login_required(role=["admin", "hr"])
def attendance():
    if request.method == "POST":
        emp_id = request.form.get("emp_id")
        att_date = request.form.get("att_date")
        status = request.form.get("status")
        try:
            dbmod.execute_query(
                """INSERT INTO attendance (emp_id, att_date, status) VALUES (%s,%s,%s)
                   ON DUPLICATE KEY UPDATE status=%s""",
                (emp_id, att_date, status, status), commit=True
            )
            flash("Attendance recorded.", "success")
        except Exception as e:
            flash(f"Error: {e}", "danger")
        return redirect(url_for("attendance"))

    employees = dbmod.execute_query(
        "SELECT emp_id, emp_code, first_name, last_name FROM employees WHERE status='Active'",
        fetchall=True
    )
    today_records = dbmod.execute_query(
        """SELECT a.*, e.first_name, e.last_name, e.emp_code
           FROM attendance a JOIN employees e ON a.emp_id = e.emp_id
           WHERE a.att_date = CURDATE() ORDER BY a.att_id DESC""",
        fetchall=True
    )
    return render_template("attendance.html", employees=employees, today_records=today_records,
                            today=date.today().isoformat())


# ==========================================================================
# REPORTS / EXPORTS
# ==========================================================================

@app.route("/reports")
@login_required()
def reports():
    dept_stats = dbmod.execute_query("SELECT * FROM department_stats", fetchall=True)
    salary_bands = dbmod.execute_query(
        """SELECT
             CASE
               WHEN basic_salary < 30000 THEN 'Below 30K'
               WHEN basic_salary BETWEEN 30000 AND 50000 THEN '30K - 50K'
               WHEN basic_salary BETWEEN 50001 AND 70000 THEN '50K - 70K'
               ELSE 'Above 70K'
             END AS band,
             COUNT(*) AS c
           FROM employees WHERE status='Active'
           GROUP BY band""",
        fetchall=True
    )
    return render_template("reports.html", dept_stats=dept_stats, salary_bands=salary_bands)


@app.route("/export/excel")
@login_required()
def export_excel():
    employees = dbmod.execute_query("SELECT * FROM employee_full_view ORDER BY emp_id", fetchall=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = ["Emp Code", "Full Name", "Email", "Phone", "Department",
               "Designation", "Date of Joining", "Basic Salary", "Allowance",
               "Deduction", "Net Salary", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="2E5BFF", end_color="2E5BFF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for emp in employees:
        ws.append([
            emp["emp_code"], emp["full_name"], emp["email"], emp["phone"],
            emp["dept_name"], emp["designation"],
            emp["date_of_joining"].isoformat() if emp["date_of_joining"] else "",
            float(emp["basic_salary"] or 0), float(emp["allowance"] or 0),
            float(emp["deduction"] or 0), float(emp["net_salary"] or 0), emp["status"],
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    log_activity("Exported employee list to Excel.")

    return send_file(
        buffer, as_attachment=True,
        download_name=f"employees_export_{date.today().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/pdf")
@login_required()
def export_pdf():
    employees = dbmod.execute_query("SELECT * FROM employee_full_view ORDER BY emp_id", fetchall=True)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], alignment=1)

    elements = [
        Paragraph("Employee Database Management System", title_style),
        Paragraph("Employee Master Report", styles["Heading2"]),
        Spacer(1, 10),
    ]

    data = [["Code", "Name", "Department", "Designation", "Net Salary", "Status"]]
    for emp in employees:
        data.append([
            emp["emp_code"], emp["full_name"], emp["dept_name"] or "-",
            emp["designation"] or "-", f"{float(emp['net_salary'] or 0):,.0f}", emp["status"],
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5BFF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F4FF")]),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    log_activity("Exported employee list to PDF.")

    return send_file(buffer, as_attachment=True,
                      download_name=f"employees_report_{date.today().isoformat()}.pdf",
                      mimetype="application/pdf")


@app.route("/salary-slip/<int:emp_id>")
@login_required()
def salary_slip(emp_id):
    emp = dbmod.execute_query(
        "SELECT * FROM employee_full_view WHERE emp_id=%s", (emp_id,), fetchone=True
    )
    if not emp:
        abort(404)

    month_name = request.args.get("month", date.today().strftime("%B"))
    year = request.args.get("year", date.today().year)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("EMPLOYEE DATABASE MANAGEMENT SYSTEM", styles["Heading1"]),
        Paragraph(f"Salary Slip - {month_name} {year}", styles["Heading2"]),
        Spacer(1, 12),
    ]

    info_data = [
        ["Employee Code", emp["emp_code"], "Department", emp["dept_name"] or "-"],
        ["Name", emp["full_name"], "Designation", emp["designation"] or "-"],
        ["Date of Joining", str(emp["date_of_joining"] or "-"), "Status", emp["status"]],
    ]
    info_table = Table(info_data, colWidths=[100, 150, 100, 150])
    info_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F2F4FF")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F2F4FF")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 16))

    basic = float(emp["basic_salary"] or 0)
    allowance = float(emp["allowance"] or 0)
    deduction = float(emp["deduction"] or 0)
    net = basic + allowance - deduction

    salary_data = [
        ["Earnings", "Amount (₹)", "Deductions", "Amount (₹)"],
        ["Basic Salary", f"{basic:,.2f}", "Total Deductions", f"{deduction:,.2f}"],
        ["Allowances", f"{allowance:,.2f}", "", ""],
        ["Gross Total", f"{basic + allowance:,.2f}", "", ""],
    ]
    salary_table = Table(salary_data, colWidths=[125, 100, 125, 100])
    salary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5BFF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(salary_table)
    elements.append(Spacer(1, 16))
    elements.append(Paragraph(f"<b>Net Payable Salary: ₹ {net:,.2f}</b>", styles["Heading3"]))
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("This is a system-generated payslip and does not require a signature.",
                               styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)

    dbmod.execute_query(
        """INSERT INTO salary_history (emp_id, month, year, basic_salary, allowance, deduction, net_salary)
           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
        (emp_id, month_name, year, basic, allowance, deduction, net), commit=True
    )
    log_activity(f"Generated salary slip for employee ID {emp_id}.")

    return send_file(buffer, as_attachment=True,
                      download_name=f"payslip_{emp['emp_code']}_{month_name}{year}.pdf",
                      mimetype="application/pdf")


# ==========================================================================
# API endpoints (for AJAX / Chart.js)
# ==========================================================================

@app.route("/api/dept-chart-data")
@login_required()
def api_dept_chart_data():
    rows = dbmod.execute_query(
        "SELECT dept_name, total_employees FROM department_stats", fetchall=True
    )
    return jsonify({
        "labels": [r["dept_name"] for r in rows],
        "data": [r["total_employees"] for r in rows],
    })


@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
